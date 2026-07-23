import chardet
import configparser
import csv
import datetime
import glob
import gspread
import openpyxl
import os
import pandas
import re
import shutil
import time
import pydata_google_auth
from dateutil.relativedelta import relativedelta
from google.oauth2 import service_account
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

def _resolve_path(filename):
    """core.pyと同階層 / カレントディレクトリ / その1つ上、の順で実在するパスを探す。
    パッケージ化(site-packages配置)されても設定・関連ファイルを見失わないための共通ヘルパー。"""
    candidates = []
    base_file = os.path.dirname(os.path.abspath(__file__))
    candidates.append(os.path.normpath(os.path.join(base_file, filename)))
    candidates.append(os.path.normpath(os.path.join(os.getcwd(), filename)))
    candidates.append(os.path.normpath(os.path.join(os.getcwd(), "../", filename)))

    for c in candidates:
        if os.path.exists(c):
            return c
    return candidates[0]


SEIKYUDL_CONFIG_FILE = 'seikyudl_config.ini'
_config_path = _resolve_path(SEIKYUDL_CONFIG_FILE)
_config_ini = configparser.ConfigParser()
_read_files = _config_ini.read(_config_path, encoding='UTF-8')
if not _read_files:
    raise FileNotFoundError(f"seikyudl_config.ini が見つかりません: {_config_path}")

# フォルダパス類（外出し設定：[Paths]セクション）
shimokawa_pdf_outpath = _config_ini['Paths']['SHIMOKAWA_PDF_OUTPATH']
shimokawa_csv_outpath = _config_ini['Paths']['SHIMOKAWA_CSV_OUTPATH']

nishijima_pdf_outpath = _config_ini['Paths']['NISHIJIMA_PDF_OUTPATH']
nishijima_csv_outpath = _config_ini['Paths']['NISHIJIMA_CSV_OUTPATH']

SEIKYU_EXCEL_PATH = _config_ini['Paths']['SEIKYU_EXCEL_PATH']

SEIKYU_OUT_BASEPATH = _config_ini['Paths']['SEIKYU_OUT_BASEPATH']
OMRON_OUT_BASEPATH = _config_ini['Paths']['OMRON_OUT_BASEPATH']

pdf_outpath = _config_ini['Paths']['PDF_OUTPATH']
csv_outpath = _config_ini['Paths']['CSV_OUTPATH']

KAKUNOU_SV_DIR = _config_ini['Paths']['KAKUNOU_SV_DIR']
KAKUNOU_WIN_DIR1 = _config_ini['Paths']['KAKUNOU_WIN_DIR1']
KAKUNOU_WIN_DIR2 = _config_ini['Paths']['KAKUNOU_WIN_DIR2']

kinkyu_inifileName = _config_ini['Paths']['KINKYU_INIFILENAME']

G_DATE_DIR = _config_ini['Paths']['G_DATE_DIR']

DATE_NOTHING = datetime.datetime.strptime('1999/7/1', '%Y/%m/%d')

_latched_taisho = False

# function
# downloadFile


def has_any_success_record(sheet_title):
    """このシートで過去に一度でも成功実績(.dateファイル)があるか"""
    date_dir = ".//..//date//"
    if not os.path.isdir(date_dir):
        return False
    for fname in os.listdir(date_dir):
        if fname.startswith(sheet_title) and fname.endswith(".date"):
            return True
    return False

def getSEIKYU_OUT_BASEPATH_Temp(gyoshamei):

    if "オムロン検収結果書" in gyoshamei:
        return OMRON_OUT_BASEPATH
    else:
        return SEIKYU_OUT_BASEPATH

def getPDFpath(sheetname):

    if "オムロン検収結果書" in sheetname:
        return pdf_outpath
    else:
        return ""

def getCSVpath(sheetname):

    if "オムロン検収結果書" in sheetname:
        return csv_outpath
    else:
        return csv_outpath


def getSEIKYU_OUT_BASEPATH():
    return SEIKYU_OUT_BASEPATH

def reset_taisho():
    global _latched_taisho
    _latched_taisho = False

def makeExOption(tmp_download_dir):
    options = webdriver.ChromeOptions()
    #options = Options()
    #options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("start-maximized")
    options.add_argument("disable-infobars")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--remote-debugging-pipe')
    #options.add_argument('--user-agent={}'.format(random.choice(list(self.user_agents))))

    prefs = {'download.default_directory': tmp_download_dir}
    options.add_experimental_option('prefs', prefs)

    #TODO
    options.add_argument('--disable-blink-features=AutomationControlled')
    #TODO  e
    prefs = {'profile.default_content_setting_values.notifications': 2,
             'profile.default_content_setting_values.automatic_downloads': 1,  # ← これを追加
             'download.default_directory': tmp_download_dir,
             'credentials_enable_service': False,
             'download.directory_upgrade': True,
             'download.extensions_to_open': '',
             'download.prompt_for_download': False,
             "plugins.always_open_pdf_externally": True,
             'safebrowsing.enabled': True,
             "protocol_handler.excluded_schemes": {"afp": True, "data": True, "disk": True, "disks": True, "file": True,
                                                   "hcp": True, "intent": True, "itms-appss": True, "itms-apps": True,
                                                   "itms": True, "market": True, "javascript": True, "mailto": True,
                                                   "ms-help": True, "news": True, "nntp": True, "shell": True,
                                                   "sip": True, "snews": False, "vbscript": True, "view-source": True,
                                                   "vnd": {"ms": {"radio": True}}}}

    #prefs = {'download.prompt_for_download': False}
    options.add_experimental_option('prefs', prefs)

    return options
# function
# downloadFile
def prepareDownload():

    # カレントディレクトリの取得
    current_dir = os.getcwd()

    # 一時ダウンロードフォルダパスの設定
    tmp_download_dir = f'{current_dir}/tmpDownload'

    # 一時フォルダが存在していたら消す(前回のが残存しているかも)
    if os.path.isdir(tmp_download_dir):
        shutil.rmtree(tmp_download_dir)

    # 一時ダウンロードフォルダの作成
    os.mkdir(tmp_download_dir)

    return tmp_download_dir

def getBasePath():

    return SEIKYU_EXCEL_PATH

def getSEIKYU_OUT_BASEPATH():

    return SEIKYU_OUT_BASEPATH

def chg99(val):

    if int(val) >= 28:
        return "99"
    else:
        return val

def sengetsudojitsu():

    today = datetime.datetime.today()
    sengetu = today - relativedelta(months=1)
    #sengetu = today - relativedelta(days=-10)

    targetday = sengetu.strftime("%Y/%m/%d")
    return targetday

# function
# read ini file
def readTargetDate(filename):

    path = os.getcwd() + "/../"
    if os.path.isfile(path + kinkyu_inifileName):
        f = open(path +  kinkyu_inifileName, 'r')
        date = f.read()
        if date != "":
            print("!!! Emergency START !!!")
            return date.replace("\n", "")
    else:
        f = open(path + kinkyu_inifileName, 'w')

    if os.path.isfile(path + filename + '.ini'):
        f = open(path + filename + '.ini', 'r')
        date = f.read().replace("\n", "")
    else:
        f = open(path + filename + '.ini', 'w')
        date = sengetsudojitsu()
        f.write(date)
    f.close()

    return date

# function
# arite ini file
def writeTargetDate(date, filename):
    path = os.getcwd() + "/../"
    f = open(path + filename + '.date', 'w')
    f.write(date)
    f.close()

# functionte
# take Slashes
def kurikku(driver, element):

    #driver.execute_script("arguments[0].scrollIntoView(false);", element)
    element.location_once_scrolled_into_view
    driver.execute_script("arguments[0].click();", element)

# function
# take Slashes
def getSEIKYU_OUT_BASEPATH(p_date):
    list_datadate = p_date.split("/")
    return str(list_datadate[0] + list_datadate[1] + list_datadate[2])

# function
# take Slashes
def takeSlash(p_date):
    list_datadate = p_date.split("/")
    return str(list_datadate[0]).zfill(2) + str(list_datadate[1] ).zfill(2)+ str(list_datadate[2]).zfill(2)


def chkFolderNishijima(SEIKYU_OUT_BASEPATH, gyoshamei, foldermei):

    rtn = SEIKYU_OUT_BASEPATH + "//" + gyoshamei

    if foldermei != "":
        rtn = rtn + "//" + foldermei
    if not os.path.exists(rtn):
        os.makedirs(rtn)
    return rtn

def detect_encoding(path: str, sample_bytes: int = 32_000) -> str:
    p = Path(path)
    # 1) BOMチェック（高速）
    with p.open("rb") as f:
        head = f.read(4)
    if head.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if head.startswith(b"\xff\xfe"):
        return "utf-16le"
    if head.startswith(b"\xfe\xff"):
        return "utf-16be"

    # 2) サイズに応じて全体 or 一部だけ読む
    size = p.stat().st_size
    with p.open("rb") as f:
        data = f.read() if size < 1_000_000 else f.read(sample_bytes)

    # 3) 代表的な日本語エンコーディングを先に試す
    for enc in ("cp932", "shift_jis", "utf-8"):
        try:
            data.decode(enc)
            return enc
        except UnicodeDecodeError:
            pass

    # 4) 最後に chardet で推定
    result = chardet.detect(data)
    enc = result.get("encoding")
    return enc if enc else "cp932"  # 不明なら日本語CSV定番のcp932にフォールバック

def copyNishijima(oldf, SEIKYU_OUT_BASEPATH, gyosham, nakapath, newf, mae, ato, kakuchoshi):

    oldFile = oldf
    nowFile = chkFolderNishijima(SEIKYU_OUT_BASEPATH, gyosham, nakapath) + "//" + mae + newf + ato

    if os.path.exists(nowFile):
        os.remove(nowFile)
    if kakuchoshi == ".pdf":
        shutil.copyfile(oldFile, nowFile + kakuchoshi)
    if kakuchoshi == ".csv":

        enc = detect_encoding(oldf)
        print(f"判定したエンコーディング: {enc}")
        df = pandas.read_csv(oldf, encoding=enc)

        # Excelに保存（xlsx形式）
        df.to_excel(nowFile + ".xlsx", index=False, engine="openpyxl")

def convertSVPath(win_path):
    s = win_path.strip()
    # 大文字小文字を無視して "\\filesv\OTTSHARE" を Linux パスに置換
    s = re.sub(r'^[\\/]{2}filesv[\\/]ottshare', KAKUNOU_SV_DIR, s, flags=re.IGNORECASE)

    # 3) 残りの区切りを '/' に
    s = s.replace('\\', '/')

    # 4) 連続スラッシュを 1 本に（先頭含む）
    s = re.sub(r'/+', '/', s)

    # 5) 先頭に必ず '/' を付ける
    if not s.startswith('/'):
        s = '/' + s


    return s

def rename(oldf, sheetName, outPath, newf, Flg=True):
    t0 = time.time()
    oldFile = oldf
    nowFile = (chkFolder(outPath, sheetName) + "//" + newf)
    if os.path.exists(nowFile):
        os.remove(nowFile)
    if Flg:
        shutil.move(oldFile, nowFile)
    else:
        shutil.copyfile(oldFile, nowFile)
    dt = time.time() - t0
    if dt > 3:
        print(f"[SLOW rename] {dt:.1f}s : {nowFile}")

def QQQrename(oldf, sheetName, outPath, newf, Flg=True):

    oldFile = oldf
    nowFile = (chkFolder(outPath, sheetName) + "//" + newf)

    if os.path.exists(nowFile):
        os.remove(nowFile)
    if Flg:
        shutil.move(oldFile, nowFile)
    else:
        shutil.copyfile(oldFile, nowFile)

def writePdfText(pdfFileName, text):

    #print("TTTTT" + text)
    if "pdf" in pdfFileName:
        nowFiletext = pdfFileName.replace(".pdf", ".txt")
        with open(nowFiletext, "w", encoding="utf-8") as f:
            f.write(text)

def chkFolder(outPath, sheetName):

    rtn = outPath
    aaa = getTantoBusho(sheetName)

    rtn = getTantoFolder(outPath, aaa)

    if not os.path.exists(rtn):
        os.makedirs(rtn)
    return rtn

def getTantoBusho(param):

    parts = param.split('_')

    if len(parts) < 2:
        return None

    target = parts[-2]

    # 末尾の数字を削除
    target = re.sub(r'\d+$', '', target)

    return target

def getTantoFolder(base_path, keyword):
    for name in os.listdir(base_path):
        full_path = os.path.join(base_path, name)
        if os.path.isdir(full_path):
            if keyword in name:
                return base_path + name
    return base_path + "0000：" + keyword

def QQQQQchkFolder(outPath, gyoshamei, targetSheet):

    rtn = outPath
    if gyoshamei != "":
        rtn = rtn + gyoshamei
    if targetSheet != "":
        rtn = rtn + "//" + targetSheet
    if not os.path.exists(rtn):
        os.makedirs(rtn)
    return rtn

def rename2(oldf, newd, newf, Flg = True):

    oldFile = oldf
    nowFile = (chkFolder("", newd, "") + newf)

    dirname = os.path.dirname(nowFile)
    if os.path.exists(nowFile):
        os.remove(nowFile)
    if Flg:
        shutil.move(oldFile, nowFile)
        dirname = os.path.dirname(nowFile)
        print(os.listdir(dirname))
    else:
        shutil.copyfile(oldFile, nowFile)
    return True


def getwrbDriver(url=None):
    wk_dldir = prepareDownload()
    #pathx = r"/usr/bin/chromedriver"
    service = Service()
    opt = makeExOption(wk_dldir)

    if url and "sagawa-exp.co.jp" in url:
        profile_dir = "/home/seikyudl/chrome_profiles/sagawa"

        # 前回のロックファイルを削除
        #lock_file = os.path.join(profile_dir, "Default", "lockfile")
        #singleton = os.path.join(profile_dir, "SingletonLock")
        #for lf in [lock_file, singleton]:
        #    if os.path.exists(lf):
        #        os.remove(lf)
        # ★追加（Linuxクラッシュ対策）
        opt.add_argument("--no-sandbox")
        opt.add_argument("--disable-dev-shm-usage")
        opt.add_argument(f"--user-data-dir={profile_dir}")
        opt.add_argument("--profile-directory=Default")

    drvr = webdriver.Chrome(options=opt, service=service)
    return drvr, wk_dldir


def regetTable(driver, p_i, p_trs, p_tablename):

    wait = WebDriverWait(driver, 30)
    tableElem = wait.until(EC.visibility_of_element_located((By.ID, p_tablename)))
    wait = WebDriverWait(tableElem, 20)
    p_trs = wait.until(EC.visibility_of_all_elements_located((By.TAG_NAME, "tr")))
    wait = WebDriverWait(p_trs[p_i], 30)
    p_tds = wait.until(EC.visibility_of_all_elements_located((By.TAG_NAME, "td")))
    return p_tds



# function
# get Latest Download File
def QQQQQgetLatestDownloadedFileName(w_TmpDlDir):

    timeout_second = 120
    end_time = time.time() + timeout_second

    while time.time() < end_time:

        files = glob.glob(os.path.join(w_TmpDlDir, "*"))

        # .crdownload を除外
        completed_files = [
            f for f in files
            if not f.endswith(".crdownload")
        ]

        if completed_files:
            # 更新時刻が一番新しいファイル
            latest_file = max(completed_files, key=os.path.getmtime)
            size_before = os.path.getsize(latest_file)
            time.sleep(1)
            size_after = os.path.getsize(latest_file)
            if size_before == size_after:
                return os.path.dirname(latest_file), os.path.basename(latest_file)

        time.sleep(1)

    return "NOTHING", "NOTHING"

def getLatestDownloadedFileName(w_TmpDlDir, file_type):
    timeout_second = 20
    poll_interval = 0.5
    size_check_interval = 5.0

    if file_type.strip() == "*":
        pattern = os.path.join(w_TmpDlDir, "*")
    else:
        pattern = os.path.join(w_TmpDlDir, f"*.{file_type.lower()}")

    end_time = time.time() + timeout_second

    while time.time() < end_time:
        remaining = end_time - time.time()
        print(f"[TIMEOUT残り] {remaining:.1f}秒")

        files = glob.glob(pattern)
        completed = [
            f for f in files
            if not f.endswith(".crdownload")
            and not f.endswith(".tmp")
            and os.path.isfile(f)
        ]

        if completed:
            latest = max(completed, key=os.path.getmtime)
            print(f"[ファイル発見] {os.path.basename(latest)}")

            try:
                size1 = os.path.getsize(latest)
                print(f"[サイズ計測①] {size1:,} bytes")
                time.sleep(size_check_interval)
                size2 = os.path.getsize(latest)
                print(f"[サイズ計測②] {size2:,} bytes　差分: {size2 - size1:+,} bytes")

                if size1 == size2 and size1 > 0:
                    print(f"[完了] サイズ安定確認 → {os.path.basename(latest)}")
                    return os.path.dirname(latest), os.path.basename(latest)
                else:
                    print(f"[書込中] まだ増加中、次のループへ")
            except OSError:
                print(f"[警告] ファイルが消えました、次のループへ")
        else:
            print(f"[待機中] 対象ファイルなし")

        time.sleep(poll_interval)

    print(f"[TIMEOUT] {timeout_second}秒以内にファイルが確定しませんでした")
    return "NOTHING", "NOTHING"

# function
def owari():

    # 一時フォルダの削除
    current_dir = os.getcwd()
    if (os.path.isdir(f'{current_dir}\\tmpDownload') == True):
        shutil.rmtree(f'{current_dir}\\tmpDownload')

# function
# getNowday
def getNowday(hikuhi):

    dt_now = datetime.datetime.now() - datetime.timedelta(days=hikuhi)
    str_nowdate = str(dt_now.year) + "/" + str(dt_now.month) + "/" + str(dt_now.day)
    str_nowdate = str_nowdate
    return str_nowdate

def getTargetDate(filename):

    tmp_date = readTargetDate(filename)

    if tmp_date == "":
        tmp_date = "1999/07/01"
    else:
        print(" 指定日: " + tmp_date + " 以降をCsv,Pdfとも取得")
    return datetime.datetime.strptime(tmp_date.replace(" ", "").replace("\n", ""), '%Y/%m/%d')

# function
# get New FileName from table
def getTblData(g_sheet, p_inPos, p_KEY, p_outPos):

    rtn = ""
    for i in range(g_sheet.max_row, 1, -1):
        if str(g_sheet.cell(i, p_inPos).value) == p_KEY:
            rtnval = g_sheet.cell(i, p_outPos).value
            if rtnval is None:
                rtn = ""
            else:
                rtn = rtnval
            break
    return rtn

def excel2matrix(ws):

    hyo = []
    for row in ws.iter_rows(values_only=True):  # values_only=True で値だけ取得
        hyo.append(list(row))

    return hyo

def isNishijima(kokyakuno, hyo):

    rtn = [False, "", "" ,"" , ""]
    mae = ""
    ato1 = ""
    uriagesaki = ""
    for row in hyo:
        okyakuno = str(row[7]).replace(" ", "")
        if okyakuno != None:
            if okyakuno != "":
                if kokyakuno == str(row[1]):
                    if row[9] != None:
                        if str(row[9]).replace(" ", "") != "":
                            if row[25] != None:
                                mae = str(row[25])
                            if row[26] != None:
                                ato1 = str(row[26])
                            if row[6] != None:
                                uriagesaki = str(row[6])

                            return True, str(row[9]), mae, ato1, uriagesaki
    return rtn

# getExcelFileData
def getExcelFileData(ichiranfilepath):

    try:
        wb = openpyxl.load_workbook(ichiranfilepath, data_only=True)
    except:
        print("file not found !!!")
        print(ichiranfilepath)
        wb = None
    return wb

def getShimebi(targetSheet):

    rtn = ""
    sns = targetSheet.title.split("_")
    if len(sns) == 3:
        rtn = sns[2]
    return rtn

def chkShimebi(str):
    n = int(str)
    if 20 <= n <= 23:
        return "20"
    elif 30 <= n <= 33:
        return "99"
    elif 1 <= n <= 5:
        return "99"
    elif 15 <= n <= 18:
        return "15"
    return None

def getSheetdataByExcel(targetSheet):

    data = []
    for row in targetSheet.iter_rows(values_only=True):
        data.append(list(row))

    return data

def getData(p_objSheets):

    url = p_objSheets['F3'].value
    LID = p_objSheets['F5'].value
    PW = p_objSheets['F6'].value
    SID = p_objSheets['F7'].value
    #TODO
    kanriHozonPath = p_objSheets['I7'].value
    kakunousakiPath = p_objSheets['I8'].value

    return url, LID, PW, SID, kanriHozonPath, kakunousakiPath

def QQQQQQtaisho(targetSheetName):

    return targetSheetName[-2:].isdecimal()

def taisho_Kaki(gyoshamei, targetSheetName):

    print(gyoshamei + " : " + targetSheetName)

    base_dir = "./../date/"
    gyosha_txt_path = os.path.join(base_dir, gyoshamei + ".txt")

    with open(gyosha_txt_path, "a", encoding="utf-8") as f:
        base_dir = "./../date/"
        f.write(targetSheetName + "\n")

def write_okyakuno_txt(gyosha, param_str):

    with open(SEIKYU_OUT_BASEPATH + "//log//" + gyosha + ".log", 'a', encoding='utf-8') as f:
        f.write(param_str + '\n')

def taisho(gyoshamei, targetSheetName):

    print(gyoshamei + " : " + targetSheetName)

    base_dir = "./"
    if not targetSheetName[-2:].isdecimal():
        return False
    gyosha_txt_path = os.path.join(base_dir + "..//date//", gyoshamei + ".txt")

    # --- 読み込み ---
    imakoko_lines = []
    try:
        with open(gyosha_txt_path, "r", encoding="utf-8") as f:
            imakoko_lines = [l.strip() for l in f.read().splitlines() if l.strip()]
    except FileNotFoundError:
        imakoko_lines = []
    # --- 無い or 空 ---
    if not imakoko_lines:
        with open(gyosha_txt_path, "a", encoding="utf-8") as f:
            #f.write(targetSheetName + "\n")
            pass
        return True

    # --- 一致：一覧になければ追記してTrue、あればFalse ---
    if targetSheetName not in imakoko_lines:
        return True
    else:
        return False

def getBango(p_renfile, no):

    cf = open(p_renfile, "r", encoding="ms932", errors="", newline="")
    f = csv.reader(cf, delimiter=",", doublequote=True,
                   lineterminator="\r\n", quotechar='"',
                   skipinitialspace=True)

    header = next(f)
    header = next(f)
    return header[no]
    cf.close

def scrollByElemAndOffset(driver, element, offset = 0):

    driver.execute_script("arguments[0].scrollIntoView();", element)

    if (offset != 0):
        script = "window.scrollTo(0, window.pageYOffset + " + str(offset) + ");"
        driver.execute_script(script)

def Cmn_getFullPath(self, path1):

    base = os.path.dirname(os.path.abspath(__file__))
    base2 = os.path.join("../", path1)
    basex = ""
    basex = os.path.normpath(os.path.join(base, base2))
    return basex


class Cmn_SpreadIO():

    def __init__(self, config_path=None):
        super().__init__()

        #self.common = common

        # google Spread Linkage Data（設定はiniファイルに外出し／モジュール冒頭で読込済のものを再利用）
        if config_path:
            config_ini = configparser.ConfigParser()
            read_files = config_ini.read(config_path, encoding='UTF-8')
            if not read_files:
                raise FileNotFoundError(f"seikyudl_config.ini が見つかりません: {config_path}")
        else:
            config_ini = _config_ini

        self.g_json_file = config_ini['Google']['JSON_FILE']
        self.g_file_name = config_ini['Google']['FILE_NAME']
        self.g_url_Bill = config_ini['Google']['URL_BILL']
        self.scope = eval(config_ini['Google']['SCOPE'])

        self.sheet_name_list = config_ini['Sheet']['LIST']
        self.sheet_name_idpw = config_ini['Sheet']['IDPW']
        self.sheet_name_Senarios = config_ini['Sheet']['SENARIOS']

        CLIENT_ID = config_ini['Google']['CLIENT_ID']
        CLIENT_SECRET = config_ini['Google']['CLIENT_SECRET']

        scoped_credentials = pydata_google_auth.get_user_credentials(
            self.scope,
            client_id=CLIENT_ID,
            client_secret=CLIENT_SECRET
        )
        gc = gspread.authorize(scoped_credentials)

        # 共有設定したスプレッドシートキー
        SPREADSHEET_KEY = config_ini['Google']['SPREADSHEET_KEY']
        self.targetSS = self.SpIO_accessSpread()

    def SpIO_getTargetSS(self):

        return self.targetSS

    def SpIO_getFullPath(self, path1):

        return _resolve_path(path1)

    def SpIO_accessSpread(self):

        g_json_fileFull = self.SpIO_getFullPath(self.g_json_file)

        #credentials = ServiceAccountCredentials.from_json_keyfile_name(g_json_fileFull, self.scope)

        credentials = service_account.Credentials.from_service_account_file(
            g_json_fileFull,
            scopes=self.scope
        )

        gc = gspread.authorize(credentials)
        return gc.open(self.g_file_name)

    def SpIO_accessSpreadSheet(self, sheet_name):

        # sheet_list = [self.targetSS.title for ws in self.targetSS.worksheets()]
        self.wks = self.targetSS.worksheet(sheet_name)
        return self.wks

    def getDatas(self, sc):

        sc.SpIO_accessSpread
        w_sp = sc.SpIO_accessSpreadSheet(sc.sheet_name_idpw)
        str_FName = w_sp.cells("B1")
        str_URL = w_sp.cells("B5")
        str_LID = w_sp.cells("B6")
        str_PW = w_sp.cells("B7")
        str_SID = w_sp.cells("B8")
        w_sp = sc.SpIO_accessSpreadSheet(sc.sheet_name_idpw)
        return str_FName, str_URL, str_LID, str_PW, str_SID, w_sp.get_all_values()


def wareki2seireki(wareki_s, def_value=0):

    era_dic = {"明治": 1868, "大正": 1912, "昭和": 1926, "平成": 1989, "令和": 2019}
    s = re.match(r'(明治|大正|昭和|平成|令和)([0-9]+|元)年', str(wareki_s))
    if s is None: return def_value
    y = int(s.group(2)) if s.group(2) != '元' else 1
    return str(era_dic[s.group(1)] + y - 1) + "/" + wareki_s[s.regs[0][1]:].replace("月", "/").replace("日", "")

############################################################################################################################
def NNNNisApprobe(datefileName, seikyu_date_str, flg=False):
    return True

##############################################################################
def isApprobe(datefileName, seikyu_date, flg=False):
    """»èÌÝB±±ÅÍ®¹L^ð©È¢B"""
    if os.path.isfile(".//..//date//" + datefileName) == False:
        return True
    file_date_str = readDate(".//..//date//" + datefileName).replace('\n', '')
    zenkai_date = datetime.datetime.strptime(file_date_str, '%Y/%m/%d')
    if flg == False:
        return seikyu_date > zenkai_date
    return seikyu_date >= zenkai_date


def write_row_log(gyoshamei, sheet_title, row_i, shitenNo, kokyakuNo, status, msg=""):
    now = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    line = f"{now} [{status}] sheet={sheet_title} row={row_i} shiten={shitenNo} kokyaku={kokyakuNo}"
    if msg:
        line += f" : {msg}"
    write_okyakuno_txt(gyoshamei, line)


def readDate(filename):

    if os.path.isfile(filename):
        f = open(filename, 'r')
        date = f.read().replace("\n", "")
    else:
        f = open(filename, 'w')
        date = datetime.datetime.now().strftime('%Y/%m/%d')
        f.write(date)
    f.close()

    return date


# arite ini file
def writeDate(date, filename):
    f = open(".//..//date//" + filename, 'w')
    f.write(date)
    f.close()

# main
if __name__ == '__main__':
    main()
